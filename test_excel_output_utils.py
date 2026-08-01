import unittest

from openpyxl import Workbook
from openpyxl.styles import Border, PatternFill, Side

from excel_output_utils import HYPERLINK_FONT_COLOR, format_literature_worksheet


class LiteratureWorksheetFormattingTests(unittest.TestCase):
    def test_hyperlinks_preserve_existing_fill_and_border(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["DOI", "link"])
        worksheet.append(["10.1000/example", "https://example.com/article"])

        fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        border = Border(bottom=Side(style="thin", color="808080"))
        for cell in worksheet[2]:
            cell.fill = fill
            cell.border = border

        format_literature_worksheet(worksheet)

        for cell in worksheet[2]:
            self.assertEqual(cell.fill, fill)
            self.assertEqual(cell.border, border)
            self.assertIsNotNone(cell.hyperlink)
            self.assertEqual(cell.font.underline, "single")
            self.assertEqual(cell.font.color.type, "rgb")
            self.assertEqual(cell.font.color.rgb, f"00{HYPERLINK_FONT_COLOR}")

        self.assertEqual(worksheet["A2"].value, "https://doi.org/10.1000/example")


if __name__ == "__main__":
    unittest.main()
